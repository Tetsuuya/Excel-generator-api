# -*- coding: utf-8 -*-
"""
Test generator for comprehensive Hospital Analytics (1,000 records, 9 worksheets, KPI cards, charts).
"""

import os
import sys
import time
import openpyxl

from dotenv import load_dotenv
load_dotenv()

from services.excel_service import ExcelService

HOSPITAL_PROMPT = """Create a complete Hospital Operations and Patient Analytics System Excel workbook with 1,000 realistic patient visit records (Jan-Dec 2025).

Use Pandas to generate the 1,000 records and compute aggregations, then write into openpyxl using a compact helper function for styling across sheets.

Create these 9 worksheets:
1. README: Title, Purpose, and worksheet directory.
2. Raw Data: 1,000 records (Patient ID, Visit ID, Name, Age, Gender, DOB, Visit Date, Dept, Doctor, Diagnosis, Treatment, Length of Stay, Medication Cost, Room Cost, Lab Cost, Total Bill, Insurance Coverage, Patient Payment, Satisfaction 1-5, Status).
3. Patient Summary: Summary aggregated by patient.
4. Department Analysis: Department metrics (Visits, Total Revenue, Avg Bill, Avg Stay, Avg Satisfaction).
5. Doctor Performance: Doctor revenue and visit counts (Top 10 ranked).
6. Monthly Trends: Jan-Dec 2025 monthly totals and MoM growth.
7. Financial Analysis: Totals by Insurance Provider and Payment Method.
8. Diagnosis Analysis: Top diagnoses by case volume and total cost.
9. Dashboard: Executive KPI cards (Total Patients, Total Revenue, Avg Bill, Avg Satisfaction, Avg Stay) plus native openpyxl Charts (Monthly Revenue LineChart, Revenue by Dept BarChart).

Ensure clean Navy headers, zebra striping, currency formats ("$"#,##0), visible gridlines, and save to output_path.
"""

def main():
    print("=" * 65)
    print("Starting Hospital Analytics Mega-Workbook Generation (1,000 Records, 9 Sheets)...")
    print("=" * 65)
    
    service = ExcelService()
    print(f"Endpoint: {service.base_url}")
    print(f"Default Model: {service.default_model}")
    print("Sending prompt to LLM...")
    
    start = time.time()
    try:
        result = service.generate_excel(
            prompt=HOSPITAL_PROMPT,
            preferred_model=os.getenv("EXCEL_MODEL", "openrouter/free")
        )
        output_file = "hospital_analytics_output.xlsx"
        with open(output_file, "wb") as f:
            f.write(result["excel_bytes"])
            
        elapsed = round(time.time() - start, 2)
        print(f"\nSUCCESS! Generated '{output_file}' in {elapsed}s using {result['model']} (Attempts: {result['attempts']})")
        print(f"File Size: {len(result['excel_bytes']) / 1024:.1f} KB\n")
        
        # Inspect the workbook
        wb = openpyxl.load_workbook(output_file, data_only=False)
        print(f"Generated Worksheets ({len(wb.sheetnames)} total):")
        for sname in wb.sheetnames:
            ws = wb[sname]
            charts_count = len(ws._charts)
            print(f"  - [{sname}] -> {ws.max_row} rows, {ws.max_column} cols, {charts_count} embedded chart(s)")
            
    except Exception as err:
        print(f"\nGeneration failed: {err}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
