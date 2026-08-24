"""
Test Suite and Interactive CLI for the Excel Generator Service.
Runs unit tests on AST sanitizer & sandbox execution, and optionally tests live Groq prompts.
"""

import os
import sys

# Ensure UTF-8 output on Windows consoles
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

from dotenv import load_dotenv

load_dotenv()

from core.sanitizer import sanitize_python_code, SecurityError
from core.executor import execute_excel_code, ExecutionError


def test_security_sanitizer():
    print("\n--- 1. Testing AST Security Sanitizer ---")

    # Test Case 1: Unauthorized import (os)
    bad_code_1 = """
import os
def generate_excel(output_path):
    os.system("echo hacked")
"""
    try:
        sanitize_python_code(bad_code_1)
        print("❌ FAIL: Failed to catch 'import os'")
    except SecurityError as e:
        print(f"✅ PASS: Caught unauthorized import: {e}")

    # Test Case 2: Forbidden call (eval)
    bad_code_2 = """
import openpyxl
def generate_excel(output_path):
    eval("__import__('os').system('dir')")
"""
    try:
        sanitize_python_code(bad_code_2)
        print("❌ FAIL: Failed to catch 'eval'")
    except SecurityError as e:
        print(f"✅ PASS: Caught forbidden call: {e}")

    # Test Case 3: Dunder traversal (__subclasses__)
    bad_code_3 = """
import openpyxl
def generate_excel(output_path):
    x = ().__class__.__bases__[0].__subclasses__()
"""
    try:
        sanitize_python_code(bad_code_3)
        print("❌ FAIL: Failed to catch dunder traversal")
    except SecurityError as e:
        print(f"✅ PASS: Caught dunder traversal: {e}")

    # Test Case 4: Valid openpyxl script
    valid_code = """
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

def generate_excel(output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Model"
    ws.views.sheetView[0].showGridLines = True
    
    # Header
    ws.append(["Metric", "Q1", "Q2", "Q3", "Q4", "Total"])
    header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    # Data
    ws.append(["Revenue", 100000, 125000, 140000, 180000, "=SUM(B2:E2)"])
    ws.append(["COGS", 40000, 50000, 56000, 72000, "=SUM(B3:E3)"])
    ws.append(["Gross Profit", "=B2-B3", "=C2-C3", "=D2-D3", "=E2-E3", "=F2-F3"])
    
    for row in ws.iter_rows(min_row=2, max_row=4, min_col=2, max_col=6):
        for cell in row:
            cell.number_format = '"$"#,##0'
            
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)
        
    wb.save(output_path)
"""
    try:
        sanitize_python_code(valid_code)
        print("✅ PASS: Valid openpyxl code passed AST security checks.")
    except Exception as e:
        print(f"❌ FAIL: Valid code was wrongly rejected: {e}")

    return valid_code


def test_sandbox_execution(valid_code: str):
    print("\n--- 2. Testing Sandbox Execution ---")
    try:
        excel_bytes = execute_excel_code(valid_code)
        output_file = "test_output.xlsx"
        with open(output_file, "wb") as f:
            f.write(excel_bytes)
        print(f"✅ PASS: Successfully executed script and created '{output_file}' ({len(excel_bytes)} bytes)")
    except Exception as e:
        print(f"❌ FAIL: Execution failed: {e}")


def test_timeout_protection():
    print("\n--- 3. Testing Timeout Protection ---")
    infinite_loop_code = """
import openpyxl
def generate_excel(output_path: str):
    while True:
        pass
"""
    try:
        execute_excel_code(infinite_loop_code, timeout_seconds=2)
        print("❌ FAIL: Infinite loop did not time out")
    except ExecutionError as e:
        print(f"✅ PASS: Infinite loop correctly terminated by sandbox: {e}")


def test_live_prompt():
    print("\n--- 4. Testing Live LLM Generation ---")
    has_key = bool(os.getenv("DEEPSEEK_API_KEY") or os.getenv("GROQ_API_KEY") or os.getenv("OPENAI_API_KEY"))
    if not has_key:
        print("⚠️ SKIP: No API key set in .env. Skipping live LLM test.")
        return

    from services.excel_service import ExcelService
    service = ExcelService()
    prompt = "Create a 12-month marketing budget breakdown with channels (SEO, PPC, Social, Content), monthly allocations, quarterly summaries, and clean openpyxl styles."
    
    print(f"Provider Endpoint: {service.base_url}")
    print(f"Prompt: {prompt}")
    print(f"Requesting Excel code using {service.default_model}...")
    
    try:
        result = service.generate_excel(prompt=prompt)
        output_file = "live_test_output.xlsx"
        with open(output_file, "wb") as f:
            f.write(result["excel_bytes"])
        print(f"✅ PASS: Generated '{output_file}' ({len(result['excel_bytes'])} bytes) in {result['duration_seconds']}s using {result['model']}!")
    except Exception as e:
        print(f"❌ FAIL: LLM generation failed: {e}")


if __name__ == "__main__":
    print("=== Running Excel Generator Test Suite ===")
    valid_code = test_security_sanitizer()
    test_sandbox_execution(valid_code)
    test_timeout_protection()
    test_live_prompt()
    print("\n=== All Core Verification Tests Completed ===")
