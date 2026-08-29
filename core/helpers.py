"""
Executive Multi-Tab Excel Helper Toolkit.
Provides high-level, rock-solid functions to build enterprise-grade, multi-tab Excel workbooks (.xlsx)
with KPI summary cards, raw data registries, analytical breakdown sheets, and native charts.
"""

from typing import List, Tuple, Any, Optional, Dict, Union
import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

# --- Corporate Executive Palette ---
COLOR_NAVY = "0F172A"       # Primary dark header
COLOR_ICE = "F1F5F9"        # KPI Card background
COLOR_ZEBRA = "F8FAFC"      # Soft zebra striping
COLOR_BORDER = "CBD5E1"     # Subtle cell border
COLOR_MUTED = "64748B"      # Subtitles & secondary labels
COLOR_SUCCESS = "10B981"    # Positive deltas
COLOR_WHITE = "FFFFFF"      # Header text
COLOR_TOTAL = "E2E8F0"      # Total row background

# --- Number Format Tokens ---
FMT_CURRENCY = '"$"#,##0'
FMT_CURRENCY_CENTS = '"$"#,##0.00'
FMT_PERCENT = '0.0%'
FMT_NUMBER = '#,##0'
FMT_DATE = 'YYYY-MM-DD'

# --- Borders ---
THIN_SIDE = Side(style='thin', color=COLOR_BORDER)
THICK_NAVY = Side(style='medium', color=COLOR_NAVY)
DOUBLE_BOTTOM = Side(style='double', color=COLOR_NAVY)
CARD_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def create_workbook(title: str = "Executive Dashboard", subtitle: str = "Generated via AI Analytics") -> Tuple[Workbook, Worksheet]:
    """Creates a new styled workbook with metadata title banner on the primary Dashboard tab."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = True

    # Margins
    ws.row_dimensions[1].height = 12
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 18

    # Title Banner
    ws["B2"] = title
    ws["B2"].font = Font(name="Segoe UI", size=16, bold=True, color=COLOR_NAVY)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")

    if subtitle:
        ws["B3"] = subtitle
        ws["B3"].font = Font(name="Segoe UI", size=9, italic=True, color=COLOR_MUTED)
        ws["B3"].alignment = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 3
    return wb, ws


def add_kpi_cards(ws: Worksheet, cards: List[Tuple[str, Any, str]], start_col_idx: int = 2, start_row: int = 5) -> None:
    """
    Renders 3-5 side-by-side executive KPI summary cards.
    cards: List of (label, value, context_delta)
    """
    ws.row_dimensions[start_row].height = 18
    ws.row_dimensions[start_row + 1].height = 28
    ws.row_dimensions[start_row + 2].height = 18

    for i, (label, val, delta) in enumerate(cards):
        col_c1 = start_col_idx + (i * 2)
        col_c2 = col_c1 + 1
        c1_letter = get_column_letter(col_c1)
        c2_letter = get_column_letter(col_c2)

        # Style card background & border
        for r in range(start_row, start_row + 3):
            for c in range(col_c1, col_c2 + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill("solid", fgColor=COLOR_ICE)
                cell.border = CARD_BORDER

        ws.merge_cells(f"{c1_letter}{start_row}:{c2_letter}{start_row}")
        ws.merge_cells(f"{c1_letter}{start_row + 1}:{c2_letter}{start_row + 1}")
        ws.merge_cells(f"{c1_letter}{start_row + 2}:{c2_letter}{start_row + 2}")

        # Label
        lbl_cell = ws.cell(row=start_row, column=col_c1, value=label.upper())
        lbl_cell.font = Font(name="Segoe UI", size=8.5, bold=True, color=COLOR_MUTED)
        lbl_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

        # Value
        val_cell = ws.cell(row=start_row + 1, column=col_c1, value=val)
        val_cell.font = Font(name="Segoe UI", size=16, bold=True, color=COLOR_NAVY)
        val_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        if isinstance(val, (int, float)) and val >= 1000:
            val_cell.number_format = FMT_CURRENCY

        # Delta badge
        delta_cell = ws.cell(row=start_row + 2, column=col_c1, value=delta)
        delta_cell.font = Font(name="Segoe UI", size=8.5, italic=True, color=COLOR_SUCCESS if "▲" in str(delta) or "+" in str(delta) else COLOR_MUTED)
        delta_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)


def add_table(
    ws: Worksheet,
    start_row: int,
    start_col: int,
    headers: List[str],
    data: List[List[Any]],
    currency_cols: Optional[List[int]] = None,
    percent_cols: Optional[List[int]] = None,
    date_cols: Optional[List[int]] = None,
    show_total: bool = False,
    zebra: bool = True
) -> int:
    """
    Renders a styled data table with headers, zebra striping, proper formats, and optional accounting total row.
    Returns the next available row index.
    """
    currency_cols = currency_cols or []
    percent_cols = percent_cols or []
    date_cols = date_cols or []

    # 1. Header Row
    ws.row_dimensions[start_row].height = 26
    for idx, h in enumerate(headers):
        c = ws.cell(row=start_row, column=start_col + idx, value=h)
        c.fill = PatternFill("solid", fgColor=COLOR_NAVY)
        c.font = Font(name="Segoe UI", size=10, bold=True, color=COLOR_WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = Border(left=THIN_SIDE, right=THIN_SIDE, top=THICK_NAVY, bottom=THICK_NAVY)

    # 2. Data Rows
    current_row = start_row + 1
    for r_idx, row_vals in enumerate(data):
        ws.row_dimensions[current_row].height = 20
        is_even = (r_idx % 2 == 0)
        row_fill = PatternFill("solid", fgColor=COLOR_ZEBRA if (is_even and zebra) else COLOR_WHITE)

        for c_idx, val in enumerate(row_vals):
            col_num = start_col + c_idx
            c = ws.cell(row=current_row, column=col_num, value=val)
            c.font = Font(name="Segoe UI", size=10, color=COLOR_NAVY)
            c.fill = row_fill
            c.border = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

            # Formatting
            if c_idx in currency_cols:
                c.number_format = FMT_CURRENCY
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in percent_cols:
                c.number_format = FMT_PERCENT
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif c_idx in date_cols:
                c.number_format = FMT_DATE
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif isinstance(val, (int, float)):
                c.number_format = FMT_NUMBER
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")

        current_row += 1

    # 3. Total Row
    if show_total:
        ws.row_dimensions[current_row].height = 22
        total_fill = PatternFill("solid", fgColor=COLOR_TOTAL)

        lbl = ws.cell(row=current_row, column=start_col, value="TOTAL")
        lbl.font = Font(name="Segoe UI", size=10, bold=True, color=COLOR_NAVY)
        lbl.fill = total_fill
        lbl.border = Border(top=THIN_SIDE, bottom=DOUBLE_BOTTOM)

        for c_idx in range(1, len(headers)):
            col_num = start_col + c_idx
            col_let = get_column_letter(col_num)
            c = ws.cell(row=current_row, column=col_num)
            c.fill = total_fill
            c.font = Font(name="Segoe UI", size=10, bold=True, color=COLOR_NAVY)
            c.border = Border(top=THIN_SIDE, bottom=DOUBLE_BOTTOM)

            if c_idx in currency_cols or c_idx in percent_cols:
                if c_idx in percent_cols:
                    c.value = f"=AVERAGE({col_let}{start_row + 1}:{col_let}{current_row - 1})"
                    c.number_format = FMT_PERCENT
                else:
                    c.value = f"=SUM({col_let}{start_row + 1}:{col_let}{current_row - 1})"
                    c.number_format = FMT_CURRENCY
                c.alignment = Alignment(horizontal="right", vertical="center")
            elif len(data) > 0 and isinstance(data[0][c_idx], (int, float)):
                c.value = f"=SUM({col_let}{start_row + 1}:{col_let}{current_row - 1})"
                c.number_format = FMT_NUMBER
                c.alignment = Alignment(horizontal="right", vertical="center")

        current_row += 1

    autofit_columns(ws)
    return current_row


def add_sheet_from_df(
    wb: Workbook,
    sheet_title: str,
    df: pd.DataFrame,
    page_title: Optional[str] = None,
    subtitle: Optional[str] = None,
    currency_cols: Optional[List[int]] = None,
    percent_cols: Optional[List[int]] = None,
    date_cols: Optional[List[int]] = None,
    show_total: bool = True
) -> Worksheet:
    """
    Creates a new styled worksheet directly from a pandas DataFrame with title banner, navy headers,
    zebra striping, and column auto-fitting.
    """
    ws = wb.create_sheet(title=sheet_title)
    ws.sheet_view.showGridLines = True
    ws.column_dimensions["A"].width = 3

    # Title Banner
    p_title = page_title or sheet_title
    ws["B2"] = p_title
    ws["B2"].font = Font(name="Segoe UI", size=15, bold=True, color=COLOR_NAVY)
    
    if subtitle:
        ws["B3"] = subtitle
        ws["B3"].font = Font(name="Segoe UI", size=9, italic=True, color=COLOR_MUTED)

    headers = list(df.columns)
    data = df.values.tolist()

    add_table(
        ws=ws,
        start_row=5 if subtitle else 4,
        start_col=2,
        headers=headers,
        data=data,
        currency_cols=currency_cols,
        percent_cols=percent_cols,
        date_cols=date_cols,
        show_total=show_total
    )

    ws.freeze_panes = f"B{6 if subtitle else 5}"
    autofit_columns(ws)
    return ws


def add_raw_data_sheet(
    wb: Workbook,
    sheet_title: str,
    df: pd.DataFrame,
    currency_cols: Optional[List[int]] = None,
    date_cols: Optional[List[int]] = None,
    percent_cols: Optional[List[int]] = None
) -> Worksheet:
    """
    Creates a dedicated high-capacity Raw Data worksheet for 100 to 10,000+ granular records.
    """
    ws = wb.create_sheet(title=sheet_title)
    ws.sheet_view.showGridLines = True

    headers = list(df.columns)
    data = df.values.tolist()

    add_table(
        ws=ws,
        start_row=1,
        start_col=1,
        headers=headers,
        data=data,
        currency_cols=currency_cols,
        percent_cols=percent_cols,
        date_cols=date_cols,
        show_total=False,
        zebra=True
    )

    ws.freeze_panes = "A2"
    autofit_columns(ws)
    return ws


def autofit_columns(ws: Worksheet, min_width: int = 12) -> None:
    """Safe column width auto-calculation with padding."""
    for col in ws.columns:
        col_let = get_column_letter(col[0].column)
        max_len = max([len(str(c.value or '')) for c in col] + [min_width])
        ws.column_dimensions[col_let].width = min(max_len + 3, 40)

def add_bar_chart(
    ws: Worksheet,
    title: str,
    data_sheet: Worksheet,
    data_min_col: int,
    data_min_row: int,
    data_max_col: int,
    data_max_row: int,
    cats_min_col: int,
    cats_min_row: int,
    cats_max_row: int,
    position: str = "B12",
    width: int = 16,
    height: int = 10
) -> BarChart:
    """Adds a clean BarChart to the worksheet in 1 line of code."""
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = title
    chart.width = width
    chart.height = height
    data = Reference(data_sheet, min_col=data_min_col, min_row=data_min_row, max_col=data_max_col, max_row=data_max_row)
    cats = Reference(data_sheet, min_col=cats_min_col, min_row=cats_min_row, max_row=cats_max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, position)
    return chart


def add_pie_chart(
    ws: Worksheet,
    title: str,
    data_sheet: Worksheet,
    data_col: int,
    data_min_row: int,
    data_max_row: int,
    cats_col: int,
    cats_min_row: int,
    cats_max_row: int,
    position: str = "J12",
    width: int = 14,
    height: int = 10
) -> PieChart:
    """Adds a clean PieChart to the worksheet in 1 line of code."""
    chart = PieChart()
    chart.title = title
    chart.width = width
    chart.height = height
    data = Reference(data_sheet, min_col=data_col, min_row=data_min_row, max_row=data_max_row)
    cats = Reference(data_sheet, min_col=cats_col, min_row=cats_min_row, max_row=cats_max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, position)
    return chart
